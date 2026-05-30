"""
Validation script: Compare the corrected planning report logic against the XLSX source of truth.

This script:
1. Connects to the database
2. Queries inventory from facility_inventory_snapshot (the fix)
3. Queries sales from sales_orders (without double-counting returns)
4. Compares against the XLSX file's "uni stock" and "uni sale" columns
5. Produces a reconciliation summary
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from datetime import date, datetime, timedelta
from sqlalchemy import func, or_
from app.db.session import SessionLocal
from app.db.export_models import (
    FacilityInventorySnapshot,
    SalesOrderRecord,
    ShopifyMasterData,
)

EXCLUDED_ORDER_STATUSES = {
    "CANCELLED", "CANCELED", "RETURNED", "REFUNDED",
    "FAILED", "UNFULFILLABLE", "ERROR", "PENDING_VERIFICATION",
}


def safe_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def main():
    xlsx_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "garment-planning_2026-04-01_2026-04-30 (1).xlsx",
    )

    if not os.path.exists(xlsx_path):
        print(f"ERROR: XLSX file not found at {xlsx_path}")
        return

    # Load XLSX
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    print(f"XLSX: {len(data_rows)} data rows")
    print(f"Headers: {headers}")
    print()

    # Build expected maps from XLSX
    xlsx_uni_stock = {}  # sku -> uni stock
    xlsx_uni_sale = {}   # sku -> uni sale
    xlsx_good_inv = {}   # sku -> good_inventory (our current value)
    xlsx_net_sale = {}   # sku -> net_sale_qty (our current value)

    for r in data_rows:
        sku = str(r[1] or "").strip().upper()
        if not sku:
            continue
        us = safe_int(r[17])
        usale = safe_int(r[20])
        gi = safe_int(r[6])
        ns = safe_int(r[5])

        if us is not None:
            xlsx_uni_stock[sku] = us
        if usale is not None:
            xlsx_uni_sale[sku] = usale
        if gi is not None:
            xlsx_good_inv[sku] = gi
        if ns is not None:
            xlsx_net_sale[sku] = ns

    print(f"XLSX SKUs with uni_stock data: {len(xlsx_uni_stock)}")
    print(f"XLSX SKUs with uni_sale data: {len(xlsx_uni_sale)}")

    # Query database with CORRECTED logic
    db = SessionLocal()
    try:
        # CORRECTED: Good Inventory from facility_inventory_snapshot
        inv_rows = (
            db.query(
                FacilityInventorySnapshot.sku.label("sku"),
                func.coalesce(func.sum(FacilityInventorySnapshot.inventory), 0).label("good_inventory"),
            )
            .filter(
                FacilityInventorySnapshot.sku.isnot(None),
                FacilityInventorySnapshot.sku != "",
            )
            .group_by(FacilityInventorySnapshot.sku)
            .all()
        )
        db_inventory_map = {
            (row.sku or "").strip().upper(): int(row.good_inventory or 0)
            for row in inv_rows
            if row.sku
        }
        print(f"DB facility_inventory_snapshot SKUs: {len(db_inventory_map)}")

        # CORRECTED: Net Sale (no double-subtraction of returns)
        start_dt = datetime(2026, 4, 1)
        end_dt = datetime(2026, 5, 1)
        
        sales_rows = (
            db.query(
                SalesOrderRecord.sku.label("sku"),
                func.sum(SalesOrderRecord.qty).label("net_sales"),
            )
            .filter(
                SalesOrderRecord.sku.isnot(None),
                SalesOrderRecord.sku != "",
                or_(
                    SalesOrderRecord.status.is_(None),
                    SalesOrderRecord.status.notin_(EXCLUDED_ORDER_STATUSES),
                ),
                or_(
                    SalesOrderRecord.sale_order_item_status.is_(None),
                    func.upper(SalesOrderRecord.sale_order_item_status).notin_(EXCLUDED_ORDER_STATUSES),
                ),
                SalesOrderRecord.order_date >= start_dt,
                SalesOrderRecord.order_date < end_dt,
            )
            .group_by(SalesOrderRecord.sku)
            .all()
        )
        db_sales_map = {
            (row.sku or "").strip().upper(): int(row.net_sales or 0)
            for row in sales_rows
            if row.sku
        }
        print(f"DB sales_orders SKUs (April 2026): {len(db_sales_map)}")
        print()

        # === RECONCILIATION ===
        
        # 1. Good Inventory: DB (corrected) vs XLSX Uni Stock
        print("=" * 80)
        print("RECONCILIATION: Good Inventory (Corrected) vs XLSX Uni Stock")
        print("=" * 80)
        inv_match = 0
        inv_mismatch = 0
        inv_mismatch_details = []
        total_db_inv = 0
        total_uni_stock = 0

        for sku, uni_stock in xlsx_uni_stock.items():
            db_inv = db_inventory_map.get(sku, 0)
            total_db_inv += db_inv
            total_uni_stock += uni_stock
            if db_inv == uni_stock:
                inv_match += 1
            else:
                inv_mismatch += 1
                if len(inv_mismatch_details) < 20:
                    inv_mismatch_details.append(
                        f"  SKU={sku}: DB_Inv={db_inv}, UniStock={uni_stock}, diff={db_inv - uni_stock}"
                    )

        print(f"Matches: {inv_match}")
        print(f"Mismatches: {inv_mismatch}")
        print(f"Total DB Inventory: {total_db_inv}")
        print(f"Total Uni Stock: {total_uni_stock}")
        print(f"Total Diff: {total_db_inv - total_uni_stock}")
        if inv_mismatch_details:
            print(f"\nSample mismatches (first 20):")
            for d in inv_mismatch_details:
                print(d)
        print()

        # 2. Net Sale: DB (corrected) vs XLSX Uni Sale
        print("=" * 80)
        print("RECONCILIATION: Net Sale (Corrected) vs XLSX Uni Sale")
        print("=" * 80)
        sale_match = 0
        sale_mismatch = 0
        sale_mismatch_details = []
        total_db_sale = 0
        total_uni_sale = 0

        for sku, uni_sale in xlsx_uni_sale.items():
            db_sale = db_sales_map.get(sku, 0)
            total_db_sale += db_sale
            total_uni_sale += uni_sale
            if db_sale == uni_sale:
                sale_match += 1
            else:
                sale_mismatch += 1
                if len(sale_mismatch_details) < 20:
                    sale_mismatch_details.append(
                        f"  SKU={sku}: DB_Sale={db_sale}, UniSale={uni_sale}, diff={db_sale - uni_sale}"
                    )

        print(f"Matches: {sale_match}")
        print(f"Mismatches: {sale_mismatch}")
        print(f"Total DB Net Sale: {total_db_sale}")
        print(f"Total Uni Sale: {total_uni_sale}")
        print(f"Total Diff: {total_db_sale - total_uni_sale}")
        if sale_mismatch_details:
            print(f"\nSample mismatches (first 20):")
            for d in sale_mismatch_details:
                print(d)
        print()

        # 3. Compare old vs new Good Inventory (how much improvement)
        print("=" * 80)
        print("IMPROVEMENT: Old Good Inventory vs New Good Inventory vs Uni Stock")
        print("=" * 80)
        old_match = sum(1 for sku, us in xlsx_uni_stock.items() if (xlsx_good_inv.get(sku, 0) or 0) == us)
        new_match = inv_match
        print(f"Old logic matches with Uni Stock: {old_match} / {len(xlsx_uni_stock)}")
        print(f"New logic matches with Uni Stock: {new_match} / {len(xlsx_uni_stock)}")
        print(f"Improvement: +{new_match - old_match} more matches")
        print()

        # 4. Compare old vs new Net Sale
        print("=" * 80)
        print("IMPROVEMENT: Old Net Sale vs New Net Sale vs Uni Sale")
        print("=" * 80)
        old_sale_match = sum(1 for sku, us in xlsx_uni_sale.items() if (xlsx_net_sale.get(sku, 0) or 0) == us)
        new_sale_match = sale_match
        print(f"Old logic matches with Uni Sale: {old_sale_match} / {len(xlsx_uni_sale)}")
        print(f"New logic matches with Uni Sale: {new_sale_match} / {len(xlsx_uni_sale)}")
        print(f"Improvement: +{new_sale_match - old_sale_match} more matches")

    finally:
        db.close()


if __name__ == "__main__":
    main()
